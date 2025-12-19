import tkinter as tk
from datetime import datetime
from pathlib import Path
from tkinter import ttk, messagebox
from typing import Tuple

import cv2
import queue
import serial
import threading
from openpyxl import Workbook, load_workbook
from time import sleep

PORT = "COM11"  # Change this to your Arduino COM port
BAUD = 9600
CAMERA_INDEX = 0  # Default webcam index

BASE_DIR = Path(__file__).resolve().parent
PHOTO_DIR = BASE_DIR / "rfid_fotograflar"
LOG_DIR = BASE_DIR / "rfid_kayitlari"
LOG_FILE = LOG_DIR / "rfid_log.xlsx"
PHOTO_DIR.mkdir(parents=True, exist_ok=True)
LOG_DIR.mkdir(parents=True, exist_ok=True)

# Update this mapping with your card UIDs and owner names
OWNERS = {
    "23 91 8F 11": "Mehmet Can Çatık",
    "03 68 B1 0D": "Ahmet Kaya",
}


def normalize_uid(uid_line: str) -> str:
    """Format UID string consistently for dictionary lookups."""
    return " ".join(part.strip().upper() for part in uid_line.split())


def capture_photo(uid: str) -> Tuple[Path, datetime]:
    """Kameradan fotoğraf çekip üzerine UID/zaman yaz."""
    cap = cv2.VideoCapture(CAMERA_INDEX)
    if not cap.isOpened():
        raise RuntimeError("Kamera açılamadı")

    ret, frame = cap.read()
    cap.release()
    if not ret or frame is None:
        raise RuntimeError("Kare alınamadı")

    timestamp = datetime.now()
    overlay_text = f"{uid}  {timestamp.strftime('%d.%m.%Y %H:%M:%S')}"
    cv2.putText(
        frame,
        overlay_text,
        (20, 40),
        cv2.FONT_HERSHEY_SIMPLEX,
        0.8,
        (0, 255, 0),
        2,
        cv2.LINE_AA,
    )

    safe_uid = uid.replace(" ", "-")
    file_name = f"{timestamp.strftime('%Y%m%d_%H%M%S')}_{safe_uid}.jpg"
    photo_path = PHOTO_DIR / file_name
    cv2.imwrite(str(photo_path), frame)
    return photo_path, timestamp


def log_to_excel(uid: str, owner: str, timestamp: datetime, photo_path: Path) -> None:
    """Excel (xlsx) dosyasına yeni satır ekle."""
    if LOG_FILE.exists():
        wb = load_workbook(LOG_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Zaman", "Kart UID", "Kart Sahibi", "Fotoğraf"])

    ws.append(
        [
            timestamp.strftime("%d.%m.%Y %H:%M:%S"),
            uid,
            owner,
            str(photo_path),
        ]
    )
    wb.save(LOG_FILE)


class RfidApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("RFID İzleyici")
        self.queue = queue.Queue()

        frame = ttk.Frame(root, padding=16)
        frame.pack(fill="both", expand=True)

        ttk.Label(frame, text="Kart UID").grid(row=0, column=0, sticky="w")
        self.uid_var = tk.StringVar(value="Bekleniyor…")
        ttk.Label(frame, textvariable=self.uid_var, font=("Segoe UI", 16)).grid(
            row=1, column=0, sticky="w"
        )

        ttk.Label(frame, text="Kart Sahibi").grid(row=2, column=0, sticky="w", pady=(16, 0))
        self.owner_var = tk.StringVar(value="-")
        ttk.Label(frame, textvariable=self.owner_var, font=("Segoe UI", 16)).grid(
            row=3, column=0, sticky="w"
        )

        ttk.Label(frame, text="Durum").grid(row=4, column=0, sticky="w", pady=(16, 0))
        self.status_var = tk.StringVar(value="-")
        ttk.Label(frame, textvariable=self.status_var, wraplength=360).grid(
            row=5, column=0, sticky="w"
        )

        ttk.Button(frame, text="Çıkış", command=self.root.destroy).grid(
            row=6, column=0, pady=(24, 0), sticky="e"
        )

        self.root.after(100, self.process_queue)

        try:
            self.ser = serial.Serial(PORT, BAUD, timeout=1)
            self.thread = threading.Thread(target=self.serial_worker, daemon=True)
            self.thread.start()
        except serial.SerialException as exc:
            messagebox.showerror("Hata", f"Seri port açılamadı:\n{exc}")
            self.ser = None

    def serial_worker(self):
        """Background thread reading Serial data from Arduino."""
        while True:
            if not self.ser:
                return
            try:
                line = self.ser.readline().decode("utf-8", errors="ignore").strip()
                if line.startswith("Kart UID:"):
                    uid = normalize_uid(line.replace("Kart UID:", "").strip())
                    owner = OWNERS.get(uid, "Tanımsız kart")
                    status = "Fotoğraf çekilemedi"
                    try:
                        photo_path, timestamp = capture_photo(uid)
                        log_to_excel(uid, owner, timestamp, photo_path)
                        status = f"Fotoğraf kaydedildi: {photo_path.name}"
                    except Exception as exc:  # noqa: BLE001
                        status = f"Hata: {exc}"
                    self.queue.put((uid, owner, status))
            except serial.SerialException:
                self.queue.put(("Bağlantı koptu", "-", "Seri port hatası"))
                return
            sleep(0.1)

    def process_queue(self):
        """Update UI with the latest UID/owner pair."""
        while not self.queue.empty():
            uid, owner, status = self.queue.get()
            self.uid_var.set(uid)
            self.owner_var.set(owner)
            self.status_var.set(status)
        self.root.after(100, self.process_queue)


def main():
    root = tk.Tk()
    app = RfidApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
